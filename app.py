import io
import math
import re
import shutil
import subprocess
import tempfile
from copy import copy
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="Planning dynamique Excel",
    page_icon="📄",
    layout="centered",
)

st.markdown(
    """
    <style>
      .block-container {max-width: 760px; padding-top: 4rem;}
      .main-title {text-align:center; font-size:2rem; font-weight:700; margin-bottom:.45rem;}
      .sub {text-align:center; color:#777; margin-bottom:2rem;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="main-title">Dépose ton fichier Excel</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="sub">Le planning dynamique est créé automatiquement. Le fichier modèle sert uniquement à la mise en page.</div>',
    unsafe_allow_html=True,
)


MODEL_PATH = Path(__file__).with_name("modele_visuel.xlsx")
MODEL_SHEET = "Planning dynamique"

TITLE_RE = re.compile(r"CHANTIER\s*:.*?(?:-\s*)?LOT\b", re.I)
LOT_RE = re.compile(r"\bLOT\s*[:\-]?\s*(.+)$", re.I)
DURATION_RE = re.compile(r"dur[ée]e.*jour", re.I)
EFFECTIF_RE = re.compile(r"effectif", re.I)
TOTAL_RE = re.compile(r"dur[ée]e\s+totale|total", re.I)


def excel_ref(sheet_name: str, coord: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!{coord}"


def formula_text(sheet_name: str, coord: str) -> str:
    ref = excel_ref(sheet_name, coord)
    return f'=IF({ref}="","",{ref})'


def formula_value(sheet_name: str, coord: str) -> str:
    return f"={excel_ref(sheet_name, coord)}"


def is_number(value):
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        s = value.strip().replace(" ", "").replace(",", ".")
        try:
            float(s)
            return True
        except Exception:
            return False
    return False


def as_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).strip().replace(" ", "").replace(",", "."))


def cell_text(cell):
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def find_duration_headers(ws):
    hits = []
    for row in ws.iter_rows():
        for cell in row:
            txt = cell_text(cell)
            if txt and DURATION_RE.search(txt):
                hits.append((cell.row, cell.column))
    return hits


def find_titles(ws):
    titles = []
    for row in ws.iter_rows():
        for cell in row:
            txt = cell_text(cell)
            if txt and TITLE_RE.search(txt):
                titles.append((cell.row, cell.column, txt))
                break
    return titles


def nearest_title_before(titles, row):
    candidates = [t for t in titles if t[0] < row]
    return candidates[-1] if candidates else None


def detect_columns(ws, header_row, duration_col, scan_start, scan_end):
    # Effectif : explicitement ignoré.
    effectif_col = None
    for c in range(1, max(1, duration_col)):
        txt = cell_text(ws.cell(header_row, c))
        if txt and EFFECTIF_RE.search(txt):
            effectif_col = c

    excluded = {duration_col}
    if effectif_col:
        excluded.add(effectif_col)

    # La désignation est le champ texte le plus dense à gauche de la durée.
    scores = []
    for c in range(1, duration_col):
        if c in excluded:
            continue
        nonempty = 0
        total_len = 0
        for r in range(scan_start, scan_end + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                nonempty += 1
                total_len += len(v.strip())
        scores.append((total_len, nonempty, c))

    if not scores:
        return None, None

    scores.sort(reverse=True)
    task_col = scores[0][2]

    # Catégorie : colonne texte la plus proche à gauche de la désignation.
    category_col = None
    for c in range(task_col - 1, 0, -1):
        if c not in excluded:
            category_col = c
            break

    return category_col, task_col


def detect_lot_label_row(ws, header_row, first_task_row, duration_col):
    for r in range(header_row + 1, first_task_row):
        populated = []
        for c in range(1, min(duration_col, 5) + 1):
            txt = cell_text(ws.cell(r, c))
            if txt:
                populated.append((c, txt))
        if len(populated) == 1 and populated[0][0] <= 2:
            txt = populated[0][1]
            if not TOTAL_RE.search(txt):
                return r, populated[0][0]
    return None


def extract_lot_name(title_text):
    m = LOT_RE.search(title_text or "")
    return m.group(1).strip() if m else ""


def detect_blocks_in_sheet(ws):
    headers = find_duration_headers(ws)
    titles = find_titles(ws)
    blocks = []

    for idx, (header_row, duration_col) in enumerate(headers):
        next_header = headers[idx + 1][0] if idx + 1 < len(headers) else ws.max_row + 1
        next_title_rows = [t[0] for t in titles if t[0] > header_row]
        next_title = min(next_title_rows) if next_title_rows else ws.max_row + 1
        end_row = min(next_header, next_title, ws.max_row + 1) - 1

        task_rows = []
        for r in range(header_row + 1, end_row + 1):
            dur = ws.cell(r, duration_col).value
            row_text = " ".join(cell_text(ws.cell(r, c)) for c in range(1, duration_col + 1))
            if TOTAL_RE.search(row_text):
                break
            if is_number(dur) and as_number(dur) > 0:
                task_rows.append(r)

        if not task_rows:
            continue

        category_col, task_col = detect_columns(
            ws, header_row, duration_col, task_rows[0], task_rows[-1]
        )
        if task_col is None:
            continue

        # Une vraie tâche doit avoir une désignation texte.
        task_rows = [
            r for r in task_rows
            if cell_text(ws.cell(r, task_col))
        ]
        if not task_rows:
            continue

        title = nearest_title_before(titles, header_row)
        if title is None:
            # On refuse d'inventer le chantier / lot.
            continue

        title_row, title_col, title_text = title
        lot_label = detect_lot_label_row(ws, header_row, task_rows[0], duration_col)

        blocks.append(
            {
                "sheet": ws.title,
                "title_row": title_row,
                "title_col": title_col,
                "title_text": title_text,
                "lot_name": extract_lot_name(title_text),
                "lot_label": lot_label,
                "category_col": category_col,
                "task_col": task_col,
                "duration_col": duration_col,
                "task_rows": task_rows,
            }
        )

    return blocks


def detect_all_blocks(wb):
    found = []
    for ws in wb.worksheets:
        # Ne jamais prendre un ancien planning dynamique comme source de données.
        if ws.title.lower().startswith("planning dynamique"):
            continue
        found.extend(detect_blocks_in_sheet(ws))
    return found


def next_planning_name(wb):
    base = "Planning dynamique"
    if base not in wb.sheetnames:
        return base
    i = 2
    while f"{base} {i}" in wb.sheetnames:
        i += 1
    return f"{base} {i}"


def copy_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def copy_row_style(template_ws, src_row, target_ws, target_row, end_col):
    for col in range(1, end_col + 1):
        if col <= 3:
            src_col = col
        else:
            # Répète exactement le motif d'une semaine du modèle (D:M = 10 demi-journées).
            src_col = 4 + ((col - 4) % 10)
        copy_style(template_ws.cell(src_row, src_col), target_ws.cell(target_row, col))

    src_height = template_ws.row_dimensions[src_row].height
    if src_height:
        target_ws.row_dimensions[target_row].height = src_height


def set_column_layout(template_ws, target_ws, end_col):
    for col in range(1, end_col + 1):
        letter = get_column_letter(col)
        if col <= 3:
            src_letter = get_column_letter(col)
        else:
            src_letter = get_column_letter(4 + ((col - 4) % 10))
        src_dim = template_ws.column_dimensions[src_letter]
        if src_dim.width is not None:
            target_ws.column_dimensions[letter].width = src_dim.width


def merge_category_groups(ws_out, out_task_rows, source_ws, source_task_rows, category_col):
    if category_col is None:
        return

    groups = []
    start_idx = 0
    for i in range(1, len(source_task_rows)):
        current_source = cell_text(source_ws.cell(source_task_rows[i], category_col))
        if current_source:
            groups.append((start_idx, i - 1))
            start_idx = i
    groups.append((start_idx, len(source_task_rows) - 1))

    for a, b in groups:
        source_r = source_task_rows[a]
        out_a = out_task_rows[a]
        out_b = out_task_rows[b]
        coord = source_ws.cell(source_r, category_col).coordinate

        if out_b > out_a:
            ws_out.merge_cells(start_row=out_a, start_column=1, end_row=out_b, end_column=1)
        ws_out.cell(out_a, 1).value = formula_text(source_ws.title, coord)


def build_planning(source_wb, template_wb):
    blocks = detect_all_blocks(source_wb)
    if not blocks:
        raise ValueError(
            "Je n'ai pas pu détecter automatiquement de bloc de planning. "
            "Le fichier doit contenir un titre CHANTIER / LOT et une colonne 'Durée ... jours'."
        )

    template_ws = template_wb[MODEL_SHEET]
    out_name = next_planning_name(source_wb)
    ws = source_wb.create_sheet(out_name)

    # Nombre de semaines : minimum 7, dimensionné selon le lot le plus long.
    max_total_days = 0.0
    for block in blocks:
        src = source_wb[block["sheet"]]
        total = sum(as_number(src.cell(r, block["duration_col"]).value) for r in block["task_rows"])
        max_total_days = max(max_total_days, total)

    weeks = max(7, int(math.ceil(max_total_days / 5.0)))
    half_day_cols = weeks * 10
    calendar_start_col = 4
    calendar_end_col = calendar_start_col + half_day_cols - 1
    helper_start_col = calendar_end_col + 1
    helper_end_col = calendar_end_col + 2

    set_column_layout(template_ws, ws, calendar_end_col)

    # Colonnes helper cachées.
    ws.column_dimensions[get_column_letter(helper_start_col)].hidden = True
    ws.column_dimensions[get_column_letter(helper_end_col)].hidden = True

    blue_fill = PatternFill("solid", fgColor="00A9E0")
    current_row = 1

    for block_index, block in enumerate(blocks):
        src = source_wb[block["sheet"]]
        task_rows_src = block["task_rows"]

        title_row = current_row
        blank1 = current_row + 1
        blank2 = current_row + 2
        week_row = current_row + 3
        first_task_out = current_row + 4
        task_rows_out = list(range(first_task_out, first_task_out + len(task_rows_src)))

        # Styles tirés du modèle uniquement.
        copy_row_style(template_ws, 1, ws, title_row, calendar_end_col)
        copy_row_style(template_ws, 2, ws, blank1, calendar_end_col)
        copy_row_style(template_ws, 3, ws, blank2, calendar_end_col)
        copy_row_style(template_ws, 4, ws, week_row, calendar_end_col)

        # Titre : référence dynamique vers la source.
        ws.merge_cells(
            start_row=title_row, start_column=1,
            end_row=title_row, end_column=calendar_end_col
        )
        title_coord = src.cell(block["title_row"], block["title_col"]).coordinate
        ws.cell(title_row, 1).value = formula_value(src.title, title_coord)

        # Nom du lot / libellé trouvé dans la source.
        if block["lot_label"]:
            lr, lc = block["lot_label"]
            ws.cell(week_row, 1).value = formula_text(src.title, src.cell(lr, lc).coordinate)
        elif block["lot_name"]:
            # Le texte provient mot pour mot du titre source.
            ws.cell(week_row, 1).value = block["lot_name"]

        # Semaines : 10 demi-colonnes par semaine.
        for w in range(weeks):
            c1 = calendar_start_col + w * 10
            c2 = c1 + 9
            ws.merge_cells(start_row=week_row, start_column=c1, end_row=week_row, end_column=c2)
            ws.cell(week_row, c1).value = f"Semaine {w + 1}"

        # Lignes de tâches.
        for i, (src_r, out_r) in enumerate(zip(task_rows_src, task_rows_out)):
            model_task_row = 5 if i == 0 else 6
            copy_row_style(template_ws, model_task_row, ws, out_r, calendar_end_col)

            task_coord = src.cell(src_r, block["task_col"]).coordinate
            dur_coord = src.cell(src_r, block["duration_col"]).coordinate

            ws.cell(out_r, 2).value = formula_text(src.title, task_coord)
            ws.cell(out_r, 3).value = formula_value(src.title, dur_coord)
            ws.cell(out_r, 3).number_format = '0.0 "j"'

            # Hauteur adaptée aux longues désignations.
            task_len = len(cell_text(src.cell(src_r, block["task_col"])))
            if task_len > 90:
                ws.row_dimensions[out_r].height = 55
            elif task_len > 55:
                ws.row_dimensions[out_r].height = 45
            else:
                ws.row_dimensions[out_r].height = max(ws.row_dimensions[out_r].height or 0, 40)

            # Helpers dynamiques en demi-journées.
            if i == 0:
                ws.cell(out_r, helper_start_col).value = "=0"
            else:
                ws.cell(out_r, helper_start_col).value = f"={get_column_letter(helper_end_col)}{out_r-1}"
            ws.cell(out_r, helper_end_col).value = (
                f"={get_column_letter(helper_start_col)}{out_r}+C{out_r}*2"
            )

            # Barres dynamiques.
            helper_start = f"${get_column_letter(helper_start_col)}{out_r}"
            helper_end = f"${get_column_letter(helper_end_col)}{out_r}"
            for c in range(calendar_start_col, calendar_end_col + 1):
                offset = c - calendar_start_col + 1
                cell = ws.cell(out_r, c)
                cell.value = f'=IF(AND({offset}>{helper_start},{offset}<={helper_end}),1,"")'
                cell.number_format = ";;;"

        # Catégories regroupées visuellement.
        merge_category_groups(
            ws, task_rows_out, src, task_rows_src, block["category_col"]
        )

        # Mise en forme conditionnelle des barres.
        cal_range = (
            f"{get_column_letter(calendar_start_col)}{first_task_out}:"
            f"{get_column_letter(calendar_end_col)}{task_rows_out[-1]}"
        )
        ws.conditional_formatting.add(
            cal_range,
            CellIsRule(operator="equal", formula=["1"], fill=blue_fill)
        )

        current_row = task_rows_out[-1] + 4

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D5"

    # Force Excel à recalculer les formules à l'ouverture.
    try:
        source_wb.calculation.calcMode = "auto"
        source_wb.calculation.fullCalcOnLoad = True
        source_wb.calculation.forceFullCalc = True
    except Exception:
        pass

    return out_name


def convert_xls_to_xlsx(file_bytes: bytes, original_name: str) -> bytes:
    """
    Conversion transparente .xls -> .xlsx côté serveur.
    LibreOffice est installé via packages.txt sur Streamlit Community Cloud.
    """
    libreoffice = shutil.which("libreoffice") or shutil.which("soffice")
    if not libreoffice:
        raise RuntimeError(
            "Le convertisseur Excel n'est pas disponible sur le serveur. "
            "Vérifie que le fichier packages.txt est bien présent sur GitHub."
        )

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        safe_name = Path(original_name).name
        if not safe_name.lower().endswith(".xls"):
            safe_name += ".xls"

        source_path = tmp_path / safe_name
        source_path.write_bytes(file_bytes)

        result = subprocess.run(
            [
                libreoffice,
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(tmp_path),
                str(source_path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=90,
        )

        output_path = tmp_path / (source_path.stem + ".xlsx")

        if result.returncode != 0 or not output_path.exists():
            details = (result.stderr or result.stdout or "").strip()
            raise RuntimeError(
                "La conversion automatique du fichier .xls a échoué."
                + (f" Détail : {details}" if details else "")
            )

        return output_path.read_bytes()


def process_file(file_bytes: bytes, original_name: str) -> bytes:
    if not MODEL_PATH.exists():
        raise FileNotFoundError("Le fichier modele_visuel.xlsx est absent du projet.")

    # L'utilisateur n'a rien à convertir : le site le fait automatiquement.
    if original_name.lower().endswith(".xls") and not original_name.lower().endswith(".xlsx"):
        file_bytes = convert_xls_to_xlsx(file_bytes, original_name)

    source_stream = io.BytesIO(file_bytes)
    source_wb = load_workbook(source_stream, data_only=False)
    template_wb = load_workbook(MODEL_PATH, data_only=False)

    build_planning(source_wb, template_wb)

    output = io.BytesIO()
    source_wb.save(output)
    output.seek(0)
    return output.getvalue()


uploaded = st.file_uploader(
    "Dépose ton document ici",
    type=["xls", "xlsx"],
    label_visibility="collapsed",
)

if uploaded is not None:
    try:
        with st.spinner("Création du planning dynamique..."):
            result = process_file(uploaded.getvalue(), uploaded.name)

        lower_name = uploaded.name.lower()
        if lower_name.endswith(".xlsx"):
            output_name = uploaded.name[:-5] + "_planning.xlsx"
        elif lower_name.endswith(".xls"):
            output_name = uploaded.name[:-4] + "_planning.xlsx"
        else:
            output_name = "fichier_planning.xlsx"

        st.success("Planning créé avec succès.")
        st.download_button(
            "Récupérer le fichier modifié",
            data=result,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as exc:
        st.error(str(exc))
